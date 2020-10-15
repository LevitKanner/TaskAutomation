"""
excel_to_csv.py - Converts all excel files in the current directory to csv files.
"""
import openpyxl as opx, csv, os 
from pathlib import Path

# Get all xlsx files
excelFiles = Path.cwd().glob("*.xlsx")

#load each file using openpyxl
for file in excelFiles: 
    wb = opx.load_workbook(file)
    sheets = wb.sheetnames
        
    for sheet in sheets:
        s = wb[sheet]
        data = map(lambda x: list(x), s.values)
        
        with open(f'{os.path.basename(file).strip(".xlsx")}_ {sheet}.csv', "w") as cs:
            destFile = csv.writer(cs)
            destFile.writerows(data)
            
            
# in each file get all sheetnames
# for each sheet get the values 
# Map data to a list of lists
# Create a new csv file with the name of the current excel file _ sheetname .csv
# write rows using csv.writeRows method
#save csv file.